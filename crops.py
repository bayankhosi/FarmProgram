from calendar import monthrange
import datetime
import openpyxl as opx
import pandas as pd
import upload
import statscalc

fmt = '%d/%m/%Y'
today = datetime.datetime.now().date()
month = datetime.datetime.now().month                           # month number
year = datetime.datetime.now().year                             # year

Crops = opx.load_workbook('./Files/crops.xlsx')

# opens current year sheet
crops = Crops.worksheets[0]
fields = Crops.worksheets[1]


class veggie():

    def planting():

        row = 1 + crops.max_row

        crop = str(input("Enter name of crop (small letters):\n\t"))
        crops.cell(row=row, column=1).value = crop

        field = str(input("Enter field number:\n\t"))
        crops.cell(row=row, column=2).value = field

        # date
        crops.cell(row=row, column=3).value = today.strftime(fmt)

        age = int(input("Enter seedling age in weeks:\n\t"))
        crops.cell(row=row, column=4).value = 7*age

    def harvest():

        field = int(input("Enter field number harvested:\n\t"))
        for i in crops.iter_rows:
            if (crops.cell(row=i, column=2).value == field) and (crops.cell(row=i, column=9).value == None):
                print('kak')


class field():

    def work():

        row = 1 + fields.max_row

        field = str(input("Enter field number:\n\t"))
        fields.cell(row=row, column=1).value = field

        fields.cell(row=row, column=2).value = today

        fertilizer = str(input("Enter fertilizer type ('0' if none):\n\t"))
        fields.cell(row=row, column=3).value = fertilizer

        till = str(input("Enter tillage type\n0. none\n1. Shallow\n2. Deep\n\t"))
        fields.cell(row=row, column=4).value = till

        mulch = str(input("0. No Mulch\n1. Mulch\n\t"))
        fields.cell(row=row, column=5).value = mulch


def main():

    choice = int(input('''
        [1] Crop Update
        [2] Field Update'''))

    if choice == 1:
        choice = int(input('''
            [1] Planting
            [2] Harvesting'''))
        if choice == 1:
            veggie.planting()
        if choice == 2:
            veggie.harvest()

    if choice == 2:
        field.work()

    Crops.save('./Files/crops.xlsx')

if __name__ == '__main__':
    main()