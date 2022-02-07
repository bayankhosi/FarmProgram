from calendar import monthrange
import datetime
from numpy import record
import openpyxl as opx
import pandas as pd
import upload
import statscalc

date_format = '%d/%m/%Y'
today = datetime.datetime.now().date()
month = datetime.datetime.now().month                           # month number
year = datetime.datetime.now().year                             # year

spread = opx.load_workbook('./Files/spread.xlsx')

# opens current year sheet
individual = spread.worksheets[0]
whole = spread.worksheets[year - 2020]
# total number of pigs
population = int(whole.cell(column=2, row=month + 1).value)
pig_id = individual['M1'].value


def buy_age(population, pig_id):                                 # recording new piglets
    print("=========================================\n\t\tPIGLET\n")

    # Number of new updates
    m_piglet = int(input("How many male piglets: "))
    f_piglet = int(input("How many female piglets: "))
    piglets = m_piglet + f_piglet

    # Population Update
    population += piglets
    whole.cell(column=2, row=month + 1).value = population
    whole.cell(column=2, row=month + 2).value = population

    bred_bought = int(input("\n[1] - Bred\n[2] - Bought\n"))

    if bred_bought == 1:  # bred

        # choose piglet parents
        sow = str(input("Mother ID: "))
        sire = str(input("Father ID: "))
        parents = sow + "," + sire

        while piglets > 0:  # used to write a new row entry for each piglet

            pig_id += 1

            # write piglet id on spreadsheet
            rw = pig_id
            individual.cell(row=rw, column=1).value = pig_id
            individual.cell(row=1, column=13).value = pig_id

            # enter date born
            date_born = today
            individual.cell(row=rw, column=2).value = date_born

            # record sire and sow
            individual.cell(row=rw, column=8).value = parents

            # sex of piglets
            if m_piglet > 0:
                m_piglet -= 1
                individual.cell(row=rw, column=12).value = 1
            else:
                individual.cell(row=rw, column=12).value = 0

            individual.cell(row=rw, column=10).value = 0    # no feed eaten
            individual.cell(row=rw, column=9).value = 0     # no meds taken
            individual.cell(row=rw, column=11).value = 0    # no purchase date
            individual.cell(row=rw, column=3).value = 0     # no purchase price

            piglets -= 1

    if bred_bought == 2:  # bred

        purchase_price = int(input("\nEnter purchase price: E"))

        age_bought = int(input("\nEnter Age of piglet (weeks): "))

        breed = str(input("\nEnter breed name (small letters): "))

        while piglets > 0:  # used to write a new row entry for each piglet

            pig_id += 1

            # write piglet id on spreadsheet
            rw = pig_id
            individual.cell(row=rw, column=1).value = pig_id
            individual.cell(row=1, column=13).value = pig_id

            # Assign Sex
            if m_piglet > 0:
                m_piglet -= 1
                individual.cell(row=rw, column=12).value = 1
            else:
                individual.cell(row=rw, column=12).value = 0

            # record purchase date
            purchase_date = today
            individual.cell(row=rw, column=11).value = purchase_date

            # calculate date born and record it
            date_born = purchase_date - datetime.timedelta(days=7 * age_bought)
            individual.cell(row=rw, column=2).value = date_born

            # record purchase price
            individual.cell(row=rw, column=3).value = purchase_price

            # record breed
            individual.cell(row=rw, column=8).value = breed

            individual.cell(row=rw, column=10).value = 0  # no feed eaten
            individual.cell(row=rw, column=9).value = 0  # no meds taken
            piglets-=1

    elif bred_bought == 3:

        pig_id += 1
        individual['M1'].value = pig_id
        rw = pig_id + 1
        print("\nThe pig's ID is: ", pig_id)

        # Assign Sex
        sex = int(input("[1] - Male\n[0] - Female\n"))
        individual.cell(row=rw, column=12).value = sex

        age_bought = int(input("\nEnter Age of piglet (weeks): "))
        purchase_date = today         # code to record date
        individual.cell(row=rw, column=11).value = purchase_date

        date_born = purchase_date - datetime.timedelta(days=7 * age_bought)
        individual.cell(row=rw, column=2).value = date_born

        purchase_price = int(input("\nEnter purchase price: E"))
        individual.cell(row=rw, column=3).value = purchase_price

        breed = str(input("\nEnter breed name (small letters): "))
        individual.cell(row=rw, column=8).value = breed

        individual.cell(row=rw, column=10).value = 0  # no feed eaten
        individual.cell(row=rw, column=9).value = 0  # no meds taken
        individual.cell(row=rw, column=13).value = pig_id

    piglets -= 1


def consumables():                                               # resources spent on well being
    print("=========================================\n\t\tCONSUMABLES\n")
    consumable_choice = int(
        input("\t[1] - Feed\n\t[2] - Medicated a pig\n\t[3] - Miscelleneous\n\t"))

    if consumable_choice == 1:
        print("\n\tEnter mass of feed bought (Kg)")
        feed_weight = int(input("\t")) + \
            whole.cell(column=3, row=month+1).value
        # record the amount
        whole.cell(column=3, row=month+1).value = feed_weight

        print("\n\tEnter price of feed bought (E)")
        feed_price = int(input("\t")) + whole.cell(column=4, row=month+1).value
        whole.cell(column=4, row=month+1).value = feed_price

        FeedPerPig = whole.cell(column=3, row=month + 1).value / \
            whole.cell(column=2, row=month + 1).value
        whole.cell(column=5, row=month + 1).value = FeedPerPig

    elif consumable_choice == 2:
        pig_id = int(input("Enter ID of pig medicated: "))
        rw = pig_id + 1
        med = float(input("Enter amount of medication (ml): "))

        individual.cell(row=rw, column=9).value += med

    elif consumable_choice == 3:
        print("\nEnter price of item (E)")
        misc_price = int(input()) + whole.cell(column=5, row=month+1).value
        whole.cell(column=5, row=month+1).value = misc_price


def sale(population):                                            # info on slaughter and sale
    print("=========================================\n\t\tSLAUGHTER\n")

    pig_id = int(input("\nEnter ID of Pig Slaughtered: "))
    rw = pig_id + 1

    # check if there is non recorded slaughter for pig_id
    if individual.cell(row=rw, column=4).value == None:

        if individual.cell(row=rw, column=1).value == None:
            print("\nSpecified ID doesn't exist\n")
            sale(population)

        # subtract from number of pigs
        population -= 1
        whole.cell(column=2, row=month + 1).value = population
        # to ensure nxt mnt pop not 0
        whole.cell(column=2, row=month + 2).value = population
        print("\nNew Population: ", population)

        # record date of slaughter
        slaughter_date = today
        individual.cell(row=rw, column=4).value = slaughter_date

        # record slaughter age
        date_born = datetime.datetime.date(
            individual.cell(row=pig_id + 1, column=2).value)
        slaughter_age = int((today - date_born).days)
        print("\nEnter Slaughter Age of pig: ", slaughter_age, "days")
        individual.cell(row=rw, column=6).value = int(slaughter_age)

        # record slauhgter mass
        slaughter_weight = float(input("\nEnter Slaughter Weight of pig: "))
        individual.cell(row=rw, column=5).value = slaughter_weight
        price_Kg = float(input("\nPrice per Kg: "))
        sale_price = slaughter_weight * price_Kg
        individual.cell(row=rw, column=7).value = sale_price

        # estimate of total food mass eaten
        purchase_date = individual.cell(row=rw, column=11).value
        month_bought = purchase_date.month - 1
        month_slaughtered = today.month

        # don't count month feed if bought after mid-month
        if purchase_date.day > 15:
            month_bought += 1

        # don't count month feed if slaughtered before mid-month
        if today.day < 15:
            month_slaughtered -= 1

        df_month = pd.read_excel('./Files/spread.xlsx',
                                 sheet_name='2021',
                                 index_col=0)
        df_month['feed_per_pig'] = df_month.feed_mass / df_month.population

        feed_eaten = df_month[month_bought: month_slaughtered]['feed_per_pig'].sum(
        )
        individual.cell(row=rw, column=10).value = feed_eaten

    else:
        print("\nThis ID is for a pig that has already been slaughtered.\nTry again.")
        sale(population)


def monitor():                                                   # view collected data
    print("=========================================\n\t\tMONITERING\n")

    View = int(
        input("""
            View data for: 
                [1] - Individual Pig   
                [2] - Whole Month Data
                [3] - Statistics 
        """))

    if View == 1:       # individual pig data
        pig_id = int(input("\nEnter ID of pig you want to view: "))

        purchase_date = datetime.datetime.date(individual.cell(
            row=pig_id + 1, column=2).value)

        date_born = datetime.datetime.date(
            individual.cell(row=pig_id + 1, column=2).value)

        print("\nDate Born: ", purchase_date)

        print("\nPurchase Price: E", individual.cell(
            row=pig_id + 1, column=3).value)

        if individual.cell(row=pig_id + 1, column=6).value == None:
            currAge = (today-date_born).days
            print("\nAge:  ", currAge, "days")
        else:
            print("\nSlaughter Age:  ", individual.cell(
                row=pig_id + 1, column=6).value, "days")
            print("\nSlaughter Weight:  ", individual.cell(
                row=pig_id + 1, column=5).value, "Kg")
            print("\nSale Price:  E", individual.cell(
                row=pig_id + 1, column=7).value)
            print("\nFeed Eaten: ", individual.cell(
                row=pig_id + 1, column=10).value)

    elif View == 2:     # month data
        month = int(input("\nEnter month number you want to view: "))

        avAge = whole.cell(column=6, row=month + 1).value

        Population = whole.cell(column=2, row=month + 1).value

        FeedMass = whole.cell(column=3, row=month + 1).value

        FeedPerPig = whole.cell(column=3, row=month + 1).value / population

        FeedPrice = whole.cell(column=4, row=month + 1).value

        print("\nData for", whole.cell(column=1, row=month + 1).value)

        print("\nPopulation:   ", Population)

        print("\nAverage Age:   ", avAge)

        print("\nFeed Mass Bought:   ", FeedMass, "Kg")

        print("\nAverage feed per pig: ", FeedPerPig, "Kg/pig")

        print("\nPrice of feed:  E", FeedPrice)

    elif View == 3:     # statistics
        graph = int(input(("""
            Choose a graph
                [1] - Mass - Age
                [2] - Month - Feed Mass
                [3] - Population - Feed Mass
        """)))
        if graph == 1:
            statscalc.stats.mass_age()

        if graph == 2:
            statscalc.stats.month_feed()

        if graph == 3:
            statscalc.stats.age_feed()


loop = 2
while loop == 2:
    action = int(input("""
    =========================================
        [1] - Record New Piglet
        [2] - Record Consumable(s)
        [3] - Record Slaughter and Sale
        [4] - View Data\n"""))

    if action == 1:

        buy_age(population, pig_id)

    elif action == 2:
        consumables()

    elif action == 3:

        sale(population)

    elif action == 4:
        monitor()

    spread.save('./Files/spread.xlsx')
    loop = int(input("\n[1] - Exit\n[2] - For Other Operation\n",))
    print("=========================================")


try:
    upload.main()
except:
    print("Couldn't upload")
