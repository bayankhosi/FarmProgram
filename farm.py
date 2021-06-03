import datetime
import openpyxl as opx
import upload

spread = opx.load_workbook('./Files/spread.xlsx')
individual = spread.worksheets[0]
whole = spread.worksheets[1]

loop = 2
today = datetime.datetime.now().date()   # date
month = int(datetime.datetime.now().strftime("%m"))  # month number
# total number of pigs
population = whole.cell(row=2, column=month + 1).value
pig_id = individual['L1'].value


def buy_age(population, pig_id):      # function for entering piglets

    population += 1     # add to number of pigs
    whole.cell(row=2, column=month + 1).value = population
    # to ensure nxt mnt pop not 0
    whole.cell(row=2, column=month + 2).value = population

    pig_id += 1         # int(input("Enter Pig ID: "))
    individual['L1'].value = pig_id
    rw = pig_id + 1
    individual.cell(row=rw, column=1).value = pig_id
    print("\nThe pig's ID is: ", pig_id)

    purchase_date = today         # code to record date



    age_bought = int(input("\nEnter Age of piglet (weeks): "))
    date_born = purchase_date - datetime.timedelta(days=7 * age_bought)
    individual.cell(row=rw, column=2).value = date_born

    purchase_price = int(input("\nEnter purchase price: "))
    individual.cell(row=rw, column=3).value = purchase_price

    # print("Population: ", population)  # , '\n', "Piglet born: ",date_born)

    return


def consumables():  # resources spent on well being
    """ Record:
            population each month
            average age each month
            feed each month """

    consumable_choice = int(
        input("\nWhich Consumable are you recording?\n1.Feed   2.Miscelleneous: "))
    if consumable_choice == 1:
        print("\nEnter mass of feed bought (Kg)")
        feed_weight = int(input()) + whole.cell(row=3, column=month+1).value
        # record the amount
        whole.cell(row=3, column=month+1).value = feed_weight

        print("\nEnter price of feed bought (E)")
        feed_price = int(input()) + whole.cell(row=4, column=month+1).value
        whole.cell(row=4, column=month+1).value = feed_price
        # av feed per pig
        # av feed per pig per pig weight
    elif consumable_choice == 2:
        print("\nEnter price of item (E)")
        misc_price = int(input()) + whole.cell(row=5, column=month+1).value
        whole.cell(row=5, column=month+1).value = misc_price


def sale(population):
    # make averages for that individual pig available
    # profit on the pig by subtracting average spend on it

    population -= 1     # add to number of pigs
    whole.cell(row=2, column=month + 1).value = population
    # to ensure nxt mnt pop not 0
    whole.cell(row=2, column=month + 2).value = population

    pig_id = int(input("\nEnter ID of Pig Slaughtered: "))
    rw = pig_id + 1

    slaughter_date = today
    individual.cell(row=rw, column=4).value = slaughter_date
    slaughter_weight = float(input("\nEnter Slaughter Weight of pig: "))



    date_born = datetime.datetime.date(
            individual.cell(row=pig_id + 1, column=2).value)
    slaughter_age = int((today - date_born).days)
    print(slaughter_age)
    individual.cell(row=rw, column=6).value = int(slaughter_age)

    individual.cell(row=rw, column=5).value = slaughter_weight
    price_Kg = float(input("\nPrice per Kg: "))
    sale_price = slaughter_weight * price_Kg
    individual.cell(row=rw, column=7).value = sale_price

    print("\nNew Population: ", population)


def monitor():
    View = int(
        input("\nView data for: \n1. Individual Pig   2. Whole Month Data : "))
    if View == 1:
        # find age
        pig_id = int(input("\nEnter ID of pig you want to view: "))

        purchase_date = datetime.datetime.date(individual.cell(
            row=pig_id + 1, column=2).value)
        
        print("\nPurchase Date: ", purchase_date)

        print("\nPurchase Price: ", individual.cell(
            row=pig_id + 1, column=3).value)

        date_born = datetime.datetime.date(
            individual.cell(row=pig_id + 1, column=2).value)
        
        print("\nAge:  ", (today-date_born).days, "days")

    elif View == 2:
        print("\nCurrent population:   ", whole.cell(
            row=2, column=month + 1).value)
        print("\nFeed Mass Bought:   ", whole.cell(
            row=3, column=month + 1).value)
        print("\nPrice of feed:  ", whole.cell(row=4, column=month + 1).value)
        print("\nTotal spent:  ", whole.cell(row=4, column=month +
              1).value + whole.cell(row=5, column=month + 1).value)


while loop == 2:
    action = int(input("""************************************************************************
            These are the operations that can be performed\n
            [1] - Record New Piglet
            [2] - Record Bought Consumable(s)
            [3] - Record Slaughter and Sale
            [4] - View Data
        """))

    if action == 1:

        buy_age(population, pig_id)

    elif action == 2:
        consumables()

    elif action == 3:
        sale(population)

    elif action == 4:
        monitor()

    spread.save('./Files/spread.xlsx')
    loop = int(input("\n1. Exit, 2. For Other Operation: ",))
    print("************************************************************************")


upload.main()