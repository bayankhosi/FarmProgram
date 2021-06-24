import openpyxl as opx
import numpy as np
import random


spread = opx.load_workbook('./Files/spread.xlsx')
individual = spread.worksheets[0]
whole = spread.worksheets[1]

age = []
weight =[]

for row in individual.iter_rows(min_row=2, max_row=1000):
    age.append(random.randrange(start=115, stop=150))
    weight.append(random.randrange(start=40, stop=60))

#print(age,weight)

for input in range(0, 999):
    individual.cell(row=input+3, column=5).value = weight[input]
    individual.cell(row=input+3, column=6).value = age[input]


spread.save('./Files/spread.xlsx')