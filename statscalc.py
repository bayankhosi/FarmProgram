import datetime
from calendar import monthrange
import openpyxl as opx
import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns
from sklearn.tree import DecisionTreeRegressor
from sklearn.preprocessing import LabelEncoder

# Dates

today = datetime.date.today()
today_dt = datetime.datetime.fromordinal(today.toordinal())
month = today.month
year = int(today.year)

# Openpyxl

spread = opx.load_workbook('./Files/spread.xlsx')
individual = spread.worksheets[0]
whole = spread.worksheets[1]

population = int(whole.cell(column=2, row=month + 1).value)
pig_id = individual['M1'].value

# DataFrames

df = pd.read_excel('./Files/spread.xlsx',                       # all pigs
                   sheet_name='individual',
                   index_col=0)

df_month = pd.read_excel('./Files/spread.xlsx',                 # monthly data
                         sheet_name='2021')

df_alive = df.loc[df.slaughter_date.isnull()].filter(           # living pigs
    ['ID', 'slaughter_weight', 'breed', 'meds', 'sex', 'feed_eaten'])
df_alive['age'] = today_dt - pd.to_datetime(df.birth_date)

# slaughtered pigs
df_slaughtered = df.loc[df.slaughter_date.isnull() == False]


class stats():

    def mass_age():         # slaughter mass - age graph

        sns.regplot(x=df_slaughtered.slaughter_age,
                    y=df_slaughtered.slaughter_weight
                    ).set_title('Mass - Age')
        plt.show()

    def month_feed():

        sns.barplot(x=df_month.month,
                    y=df_month.feed_mass
                    ).set_title('Month - Feed Mass')
        plt.show()

    def age_feed():
        sns.barplot(x=df_month.population,
                    y=df_month.feed_mass
                    ).set_title('Average Age - Feed Mass')
        plt.show()

    def average_age():      # should be done ev half of month

        month = today.month

        day = monthrange(2021, today.month)[1]//2

        mid_month = today - datetime.timedelta(days=today.day + day)

        df_alive['mid_month_age'] = (datetime.datetime.fromordinal(
            mid_month.toordinal()) - pd.to_datetime(df.birth_date))

        df_alive['mid_month_age'] = (df_alive['mid_month_age']).dt.days

        avAge = int(df_alive.mid_month_age.mean())

        whole.cell(column=6, row=month + 1).value = avAge
        # spread.save('./Files/spread.xlsx')

        return avAge

    def feed_per_pig():

        feed_p_pig = stats.average_age() / population
        whole.cell(column=5, row=month + 1).value = feed_p_pig
        spread.save('./Files/spread.xlsx')

        for id, row in df_alive.iterrows():
            cur_feed = individual.cell(row=id+1, column=10).value
            individual.cell(row=id+1, column=10).value = feed_p_pig + cur_feed
            spread.save('./Files/spread.xlsx')

        return feed_p_pig

        print(feed_p_pig)

    def optimum_age(id):
        # the use of decision tree regressor to estimate slaughter_age

        # dealing with categorical data
        cat_cols = ['breed']
        enc = LabelEncoder()
        df_slaughtered.loc[:, cat_cols] = df_slaughtered.loc[:,
                                                             cat_cols].apply(enc.fit_transform)
        df_alive.loc[:, cat_cols] = df_alive.loc[:,
                                                 cat_cols].apply(enc.fit_transform)

        # Our target variable
        y = df_slaughtered.slaughter_age

        # Our features
        features = ['slaughter_weight', 'meds', 'breed', 'sex', 'feed_eaten']
        X = df_slaughtered[features]

        # calling model
        age_model = DecisionTreeRegressor(random_state=1)

        # fitting data into model
        age_model.fit(X, y)

        df_alive['slaughter_weight'] = 55   # we aim to slaghter at 55Kg

        age_prediction = age_model.predict(
            df_alive[features].iloc[df_alive.index == id]).round(0)
        return age_prediction


# print(stats.optimum_age(id=9))

# print(df_alive)
# print(stats.average_age())

# stats.feed_per_pig()
# stats.age_feed()
